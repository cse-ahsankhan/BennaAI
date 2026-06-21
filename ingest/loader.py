import logging
from pathlib import Path
from typing import List, Dict, Any

import fitz  # PyMuPDF

logger = logging.getLogger(__name__)

try:
    from langdetect import detect, LangDetectException
    _langdetect_available = True
except ImportError:
    _langdetect_available = False
    logger.warning("langdetect not available — language detection disabled")

# Excel support (openpyxl for xlsx; xlrd for legacy xls)
try:
    import openpyxl
    _openpyxl_available = True
except ImportError:
    _openpyxl_available = False

try:
    import xlrd
    _xlrd_available = True
except ImportError:
    _xlrd_available = False

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xlam", ".xltx", ".xltm", ".xls", ".xlt"}


def _detect_language(text: str) -> str:
    """Detect language of a text snippet. Returns 'ar', 'en', or 'unknown'."""
    if not _langdetect_available or not text.strip():
        return "unknown"
    try:
        lang = detect(text)
        return lang if lang in ("ar", "en") else lang
    except LangDetectException:
        return "unknown"


def load_pdf(file_path: Path) -> List[Dict[str, Any]]:
    """
    Extract text from a PDF using PyMuPDF, one dict per page.

    Returns list of:
        {text, page_num, language, source_file}
    """
    file_path = Path(file_path)
    pages: List[Dict[str, Any]] = []

    try:
        doc = fitz.open(str(file_path))
    except Exception as exc:
        logger.error("Failed to open PDF %s: %s", file_path, exc)
        raise

    logger.info("Loading PDF: %s (%d pages)", file_path.name, len(doc))

    for page_index in range(len(doc)):
        page = doc[page_index]
        text = page.get_text("text")  # type: ignore[attr-defined]

        if not text.strip():
            logger.warning(
                "Page %d of '%s' has no extractable text — possibly scanned.",
                page_index + 1,
                file_path.name,
            )
            continue

        language = _detect_language(text[:500])

        pages.append(
            {
                "text": text,
                "page_num": page_index + 1,
                "language": language,
                "source_file": file_path.name,
            }
        )

    doc.close()
    logger.info("Loaded %d pages from '%s'", len(pages), file_path.name)
    return pages


def _sheet_to_text(sheet) -> str:
    """Convert a worksheet (openpyxl or xlrd) to a plain-text block."""
    rows = []
    if hasattr(sheet, "iter_rows"):
        # openpyxl
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = "\t".join(cells).rstrip()
            if line:
                rows.append(line)
    else:
        # xlrd
        for row_idx in range(sheet.nrows):
            cells = [str(sheet.cell(row_idx, col).value) for col in range(sheet.ncols)]
            line = "\t".join(cells).rstrip()
            if line:
                rows.append(line)
    return "\n".join(rows)


def load_excel(file_path: Path) -> List[Dict[str, Any]]:
    """
    Extract text from an Excel file — one page dict per worksheet.

    Returns list of:
        {text, page_num, language, source_file}
    """
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()
    pages: List[Dict[str, Any]] = []

    # Legacy .xls format
    if suffix == ".xls" or suffix == ".xlt":
        if not _xlrd_available:
            raise ImportError(
                "xlrd is required to read .xls files. "
                "Install it with: pip install xlrd"
            )
        wb = xlrd.open_workbook(str(file_path))
        sheet_names = wb.sheet_names()
        logger.info("Loading Excel (xlrd): %s (%d sheets)", file_path.name, len(sheet_names))
        for idx, name in enumerate(sheet_names):
            sheet = wb.sheet_by_name(name)
            text = f"Sheet: {name}\n\n" + _sheet_to_text(sheet)
            if not text.strip():
                continue
            pages.append({
                "text": text,
                "page_num": idx + 1,
                "language": _detect_language(text[:500]),
                "source_file": file_path.name,
            })
    else:
        # .xlsx and all modern formats
        if not _openpyxl_available:
            raise ImportError(
                "openpyxl is required to read .xlsx files. "
                "Install it with: pip install openpyxl"
            )
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        logger.info("Loading Excel (openpyxl): %s (%d sheets)", file_path.name, len(wb.sheetnames))
        for idx, name in enumerate(wb.sheetnames):
            sheet = wb[name]
            text = f"Sheet: {name}\n\n" + _sheet_to_text(sheet)
            if not text.strip():
                continue
            pages.append({
                "text": text,
                "page_num": idx + 1,
                "language": _detect_language(text[:500]),
                "source_file": file_path.name,
            })
        wb.close()

    logger.info("Loaded %d sheets from '%s'", len(pages), file_path.name)
    return pages


def load_document(file_path: Path) -> List[Dict[str, Any]]:
    """
    Load any supported document type — PDF or Excel.
    Dispatches to load_pdf or load_excel based on file extension.
    """
    file_path = Path(file_path)
    if file_path.suffix.lower() in EXCEL_EXTENSIONS:
        return load_excel(file_path)
    return load_pdf(file_path)
